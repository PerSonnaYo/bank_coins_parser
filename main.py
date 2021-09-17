import vk, os, time, math, re
import openpyxl as xl
from openpyxl.chart import Reference, BarChart

class Coin():
    def __init__(self, price, nominal, name, year, proof, country, metal, cost):
        self.nominal = nominal
        self.name = name
        self.year = year
        self.proof = proof
        self.country = country
        self.metal = metal
        self.price = price
        self.cost_price = cost
    def __len__(self):
        return 8
    def __getitem__(self, item):
        if item == 1:
            return self.nominal
        elif item == 2:
            return self.name
        elif item == 3:
            return self.year
        elif item == 4:
            return self.proof
        elif item == 5:
            return self.country
        elif item == 6:
            return self.metal
        elif item == 7:
            return self.cost_price
        elif item == 8:
            return self.price
        else:
            exit(1)
    def forming_country(self, c_price, c_cost):
        if self.country not in c_price:
            c_price[self.country] = 0
        c_price[self.country] += self.price
        if self.country not in c_cost:
            c_cost[self.country] = 0
        c_cost[self.country] += self.cost_price

def diagrams(col, title, name, alfa, sheet2):
    labels = Reference(sheet2, min_col=1, min_row=2, max_row=sheet2.max_row)
    data = Reference(sheet2, min_col=col, min_row=2, max_row=sheet2.max_row)
    # diagram = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=5, max_col=7)
    chart = BarChart()
    chart.type = "col"
    chart.add_data(data)
    chart.set_categories(labels)
    chart.shape = 4
    chart.title = title
    chart.style = 12
    chart.y_axis.title = name
    chart.x_axis.title = 'Countries'
    sheet2.add_chart(chart, alfa)

def launch_country(sheet, s, nb):
    nominal_val_cell = sheet.cell(1, nb)
    nominal_val_cell.value = s

def new_book(c_price, c_cost):
    try:
        open("Results.xlsx")
        wb = xl.load_workbook("Results.xlsx")
        sheet = wb['Sheet']
    except:
        wb = xl.Workbook()
        sheet = wb.active
        launch_country(sheet, 'Россия. Цена', 1)
        launch_country(sheet, 'Россия. Себестоимость', 6)
        launch_country(sheet, 'СССР. Цена', 2)
        launch_country(sheet, 'СССР. Себестоимость', 7)
        launch_country(sheet, 'Австралия. Цена', 3)
        launch_country(sheet, 'Австралия. Себестоимость', 8)
        launch_country(sheet, 'Другое. Цена', 4)
        launch_country(sheet, 'Другое. Себестоимость', 9)

    buf = {}
    row = sheet.max_row + 1
    if row < 2:
        row = 2
    for i in c_price:  # строим таблицу из себестоимости и цен для сохранения в итоговой таблице
        if i == 'Россия':
            buf[i] = [c_price[i], c_cost[i]]
            nominal_val_cell = sheet.cell(row, 1)
            nominal_val_cell.value = buf[i][0]
            nominal_val_cell1 = sheet.cell(row, 6)
            nominal_val_cell1.value = buf[i][1]
        elif i == 'СССР':
            buf[i] = [c_price[i], c_cost[i]]
            nominal_val_cell = sheet.cell(row, 2)
            nominal_val_cell.value = buf[i][0]
            nominal_val_cell1 = sheet.cell(row, 7)
            nominal_val_cell1.value = buf[i][1]
        elif i == 'Австралия':
            buf[i] = [c_price[i], c_cost[i]]
            nominal_val_cell = sheet.cell(row, 3)
            nominal_val_cell.value = buf[i][0]
            nominal_val_cell1 = sheet.cell(row, 8)
            nominal_val_cell1.value = buf[i][1]
        else:
            if 'Другое' not in buf:
                buf['Другое'] = [0 , 0]
            buf['Другое'][0] += c_price[i]
            buf['Другое'][1] += c_cost[i]

    nominal_val_cell = sheet.cell(row, 4)
    nominal_val_cell.value = buf['Другое'][0]
    nominal_val_cell1 = sheet.cell(row, 9)
    nominal_val_cell1.value = buf['Другое'][1]
    wb.save("Results.xlsx")
#------------------------------------
vk_access_token = 'xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'
session = vk.Session(access_token=vk_access_token)
vk_api = vk.API(session, v='5.102')

vk_album_id = 'xxxxxxxxx'
vk_owner_id = 'xxxxxxxxx'

g = vk_api.account.getBanned()
photos_count =  vk_api.photos.getAlbums(owner_id=vk_owner_id, album_ids=vk_album_id)
photos_count = photos_count['items'][0]['size']

counter = 0 #текущий счетчик
row = 1#счетчик загруженных монет(2 фото одна монета)
wb = xl.Workbook()
sheet = wb.create_sheet('Price')
sheet2 = wb.create_sheet('Country')
launch_country(sheet, 'Номинал', 1)
launch_country(sheet, 'Название', 2)
launch_country(sheet, 'Год', 3)
launch_country(sheet, 'Качество', 4)
launch_country(sheet, 'Страна', 5)
launch_country(sheet, 'Металл', 6)
launch_country(sheet, 'Себестоимость', 7)
launch_country(sheet, 'Цена', 8)

launch_country(sheet2, 'Страна', 1)
launch_country(sheet2, 'Себестоимость', 2)
launch_country(sheet2, 'Цена', 3)
launch_country(sheet2, 'Наценка', 4)
c_price = {}
c_cost = {}
for j in range(math.ceil(photos_count / 1000)):
    photos = vk_api.photos.get(owner_id=vk_owner_id, album_id=vk_album_id, count=1000, offset=j*1000, extended=1)
    # cost = vk_api.photos.getAllComments(owner_id=vk_owner_id, album_id=vk_album_id, count=1000, offset=j*1000)
    # cost = cost['items']
    for photo in photos['items']:
        counter += 1
        if counter % 2 == 0:
            continue #пропускаем аверс монеты
        row += 1
        comment = photo['comments']['count']
        foto_id = photo['id']
        foto_name = photo['text'].split('. ')[:7]
        coin_price = int(re.search(r'\d+', foto_name[6]).group(0))
        cost = 0
        if comment > 0:
            if row % 9 == 0:
                time.sleep(2)
            try:
                cost = vk_api.photos.getComments(owner_id=vk_owner_id, photo_id=foto_id)['items'][0]['text'] #парсим комменты
            except:
                time.sleep(10)
                cost = vk_api.photos.getComments(owner_id=vk_owner_id, photo_id=foto_id)['items'][0]['text']#вк ставит палки в колеса
            time.sleep(0.1)
            if (cost.isdigit() == False):
                cost = 0
            else:
                cost = int(cost)
        print('Загрузил фото № {} из {}'.format(row, photos_count))
        coin = Coin(coin_price, foto_name[0], foto_name[1], foto_name[2][:-1], foto_name[3], foto_name[4], foto_name[5], cost)#создаем класс с информацией о монете
        coin.forming_country(c_price, c_cost)#формируем заранее таблицу для общей статистике по странам
        for i in range(1, len(coin) + 1):
            nominal_val_cell = sheet.cell(row, i)#заполняем таблицу в экселе
            nominal_val_cell.value = coin[i]
        # if row == 10:
        # wb.save('PriceList_2021.xlsx')
        #     break
    # break

for row, i in enumerate(c_price):#строим таблицу из себестоимости и цен
    nominal_val_cell = sheet2.cell(row + 2, 1)
    nominal_val_cell.value = i
    nominal_val_cell1 = sheet2.cell(row + 2, 2)
    nominal_val_cell1.value = c_cost[i]
    nominal_val_cell2 = sheet2.cell(row + 2, 3)
    nominal_val_cell2.value = c_price[i]
    nominal_val_cell3 = sheet2.cell(row + 2, 4)
    nominal_val_cell3.value = float(1 - (c_cost[i] / c_price[i] ))

diagrams(3, 'Сумма цен по странам', "Price", 'e2', sheet2)
diagrams(2, 'Сумма себестоимостей по странам', "Cost_Price", 'e30', sheet2)
new_book(c_price, c_cost)#сохраняем итоги здесь
wb.save('PriceList_2021.xlsx')
