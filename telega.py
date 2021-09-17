import telebot
from telebot import types
import random, re
import openpyxl as xl
import string

bot = telebot.TeleBot("xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx")

@bot.message_handler(commands=['start'])
def send_welcome(message):
    bot.send_message(message.chat.id, 'Приветствую тебя user') #приветствие

    # markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    # but1 = types.KeyboardButton("Введите название")
    # but2 = types.KeyboardButton("Изменить цену")
    # markup.add(but1)

    # bot.reply_to(message, "Начинаю поиск", parse_mode='html', reply_markup=markup)
    # bot.send_sticker(message.chat.id, stic)

@bot.message_handler(content_types=["text"])#ищем по названию в таблице
def text(message):
    rb = xl.load_workbook('PriceList_2021.xlsx')
    sheet = rb['Price']
    point = False
    for row in range(2, sheet.max_row):
        cell = []
        cell.append(sheet.cell(row, 2).value)
        dano = cell[0].lower()
        find = message.text.lower()
        count = sum(1 for _ in re.finditer(r'\b%s\b' % re.escape(find), dano))
        if count == 1:
            point = True
            cell.pop(0)
            cell.append(str(sheet.cell(row, 1).value))
            cell.append(str(sheet.cell(row, 2).value))
            cell.append(str(sheet.cell(row, 3).value))
            cell.append(str(sheet.cell(row, 4).value))
            cell.append(str(sheet.cell(row, 5).value))
            cell.append(str(sheet.cell(row, 6).value))
            cell.append(str(sheet.cell(row, 7).value))
            cell.append(str(sheet.cell(row, 8).value))
            d = " ".join(cell)
            bot.send_message(message.chat.id, d)
    if point == False:
        bot.send_message(message.chat.id, 'Ничего не найдено')
    # if message.text == 'hello':
    #     bot.send_message(message.chat.id, 'И тебе hello')

# @bot.message_handler(commands=['Введите название'])
# def welcome_help(message):
#     bot.send_message(message.chat.id, 'Ищу..............')

# @bot.message_handler(func=lambda message: True)
# def menu(message):
#     if message.chat.type == 'private':
#         if message.text == "Введите название":
#             rb = xl.load_workbook('PriceList_2021.xlsx')
#             sheet = rb['Price']
#             welcome_help(message)
#             for row in range(2, sheet.max_row):
#                 if re.search(r'\bзвать\b', 'меня звать Олег, мне 35 лет'):
#                     print("Привет")

bot.polling(none_stop=True)
